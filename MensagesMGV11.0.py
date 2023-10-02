import openpyxl
import threading
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import os
from PIL import Image, ImageGrab
import time
import pyperclip

def limpiarcmd():
    os.system('cls' if os.name == 'nt' else 'clear')
 
def Bienvenido(duration):
    limpiarcmd()
    chars = ['|', '/', '-', '\\']
    num_chars = len(chars)
    start_time = time.time()
    
    while time.time() - start_time < duration:
        for i in range(num_chars):
            print(f'\rCargando {chars[i]}', end='')
            time.sleep(0.1)
    
    print('\rCarga Completa!   ')
    
    hola = """
         __  __                                  _           _____   __  __    _____ 
        |  \/  |                                (_)         / ____| |  \/  |  / ____|
        | \  / |   ___   _ __    ___    __ _     _    ___  | (___   | \  / | | (___  
        | |\/| |  / _ \ | '_ \  / __|  / _` |   | |  / _ \  \___ \  | |\/| |  \___ \ 
        | |  | | |  __/ | | | | \__ \ | (_| |   | | |  __/  ____) | | |  | |  ____) |
        |_|  |_|  \___| |_| |_| |___/  \__,_|   | |  \___| |_____/  |_|  |_| |_____/ 
                                               _/ |                                  
                                              |__/                                   
        by: JSB
        """
    print(hola)
    chars = ['⣾', '⣽', '⣻', '⢿', '⡿', '⣟', '⣯', '⣷']
    num_chars = len(chars)
    start_time = time.time()
    
    while time.time() - start_time < duration:
        for i in range(num_chars):
            print(f'\rIniciando Programa {chars[i]}', end='')
            time.sleep(0.1)
    
    print('\rIniciado Con Exito!   ')
    time.sleep(2)
    
extension1 = ["*.xlsx"]


# Programa con envio de imagenes
def program(archivo_excel, mensaje):            
    libro = openpyxl.load_workbook(archivo_excel)
    hoja = libro["Hoja1"]
    chrome_driver_path = "chromedriver.exe"  # Reemplaza con tu ruta
    chrome_service = ChromeService(chrome_driver_path)
    chrome_service.start()
    
    # Configura las opciones del navegador
    options = webdriver.ChromeOptions()
    # Configura opciones adicionales si es necesario
    # Deshabilita las notificaciones
    options.add_argument("--disable-notifications")
    # Inicializa el WebDriver utilizando el servicio y las opciones
    browser = webdriver.Chrome(service=chrome_service, options=options)

    # Abre la página web
    browser.get('https://messages.google.com/web/conversations')

    # Espera a que el elemento del QR esté presente
    qr_element = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'mw-qr-code img')))

    # Espera a que la imagen del QR cargue
    WebDriverWait(browser, 10).until(lambda driver: qr_element.get_attribute("complete"))

    # Espera a que cambie el src del elemento del QR (indicando que se escaneó)
    WebDriverWait(browser, 60).until(EC.staleness_of(qr_element))

    # Agrega tu lógica aquí para lo que quieras hacer después de escanear el QR
    # Por ejemplo, puedes imprimir un mensaje indicando que el QR se ha escaneado
    print("El QR se ha escaneado.")

    # Espera a que aparezca el botón
    WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-main-nav > div > mw-fab-link > a'))).click()

    # Buscar el elemento de nuevo mensaje
    WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-main-nav > div > mw-fab-link > a > span.mdc-button__label')))

    for fila in hoja.iter_rows(min_row=1, values_only=True):
        Celular = " - ".join(map(str, fila))
        try:
            
            # hacer click en el boton de nuevo mensaje
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-main-nav > div > mw-fab-link > a > span.mdc-button__label').click()

            # hacer click en la entrada de nuevo mensaje
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input').click()

            # agregar numero de telefono
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input').send_keys(Celular)

            # seleccionar el numero ingresado
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > div > mw-contact-selector-button > button').click()
            
            # Espera a que la casilla de texto esté presente y sea visible
            WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'mws-autosize-textarea textarea')))

            # escribir mensaje 
            browser.find_element(By.CSS_SELECTOR,'mws-autosize-textarea textarea').send_keys(mensaje)

            # enviar mensaje con enter
            browser.find_element(By.CSS_SELECTOR,'mws-autosize-textarea textarea').send_keys(Keys.ENTER)
            
            # Pega el texto en el campo de entrada usando la combinación de teclas "Control + V"
            ActionChains(browser).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            
            # Enviar imagen con Enter
            browser.find_element(By.CSS_SELECTOR, 'mws-autosize-textarea textarea').send_keys(Keys.ENTER)
                    
        except StaleElementReferenceException as e:
            print("Error: Elemento de página obsoleto. ", str(e))
        except ElementClickInterceptedException as e:
            print("Error: Intercepción de clic en el elemento. ", str(e))
        except Exception as e:
            print("Ocurrió un error inesperado: ", str(e))

    # Cierra el navegador y detiene el servicio
    browser.quit()
    chrome_service.stop()
    print("Envio completado")
        
    libro.close()

# Programa sin envio de imagenes
def programSinImagenes(archivo_excel, mensaje):            

    libro = openpyxl.load_workbook(archivo_excel)
    hoja = libro["Hoja1"]
    chrome_driver_path = "chromedriver.exe"  # Reemplaza con tu ruta
    chrome_service = ChromeService(chrome_driver_path)

    chrome_service.start()

    # Configura las opciones del navegador
    options = webdriver.ChromeOptions()
    # Configura opciones adicionales si es necesario
    # Deshabilita las notificaciones
    options.add_argument("--disable-notifications")
    # Inicializa el WebDriver utilizando el servicio y las opciones
    browser = webdriver.Chrome(service=chrome_service, options=options)

    # Abre la página web
    browser.get('https://messages.google.com/web/conversations')

    # Espera a que el elemento del QR esté presente
    qr_element = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'mw-qr-code img')))

    # Espera a que la imagen del QR cargue
    WebDriverWait(browser, 10).until(lambda driver: qr_element.get_attribute("complete"))

    # Espera a que cambie el src del elemento del QR (indicando que se escaneó)
    WebDriverWait(browser, 60).until(EC.staleness_of(qr_element))

    # Agrega tu lógica aquí para lo que quieras hacer después de escanear el QR
    # Por ejemplo, puedes imprimir un mensaje indicando que el QR se ha escaneado
    print("El QR se ha escaneado.")

    # Espera a que aparezca el botón
    WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-main-nav > div > mw-fab-link > a'))).click()

    # Buscar el elemento de nuevo mensaje
    WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-main-nav > div > mw-fab-link > a > span.mdc-button__label')))

    for fila in hoja.iter_rows(min_row=1, values_only=True):
        Celular = " - ".join(map(str, fila))
        try:
            
            # hacer click en el boton de nuevo mensaje
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-main-nav > div > mw-fab-link > a > span.mdc-button__label').click()

            # hacer click en la entrada de nuevo mensaje
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input').click()

            # agregar numero de telefono
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input').send_keys(Celular)

            # seleccionar el numero ingresado
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > div > mw-contact-selector-button > button').click()
            
            # Espera a que la casilla de texto esté presente y sea visible
            WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'mws-autosize-textarea textarea')))

            # escribir mensaje 
            browser.find_element(By.CSS_SELECTOR,'mws-autosize-textarea textarea').send_keys(mensaje)

            # enviar mensaje con enter
            browser.find_element(By.CSS_SELECTOR,'mws-autosize-textarea textarea').send_keys(Keys.ENTER)
                    
        except StaleElementReferenceException as e:
            print("Error: Elemento de página obsoleto. ", str(e))
        except ElementClickInterceptedException as e:
            print("Error: Intercepción de clic en el elemento. ", str(e))
        except Exception as e:
            print("Ocurrió un error inesperado: ", str(e))

    
    # Cierra el navegador y detiene el servicio
    browser.quit()
    chrome_service.stop()
    print("Envio completado")
        
    libro.close()

def buscar_base_de_datos():
    global BaseDatosMasiva
    print("Buscar Base de Datos")
    BaseDatosMasiva = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if BaseDatosMasiva:
        print("Base de Datos seleccionada: " + BaseDatosMasiva)

def validarMensaje():
    mensaje = str(input("El mensaje es correcto? y/n "))
    while True:
        try:
            while mensaje!="y" and mensaje!="n":
                mensaje = str(input("El mensaje es correcto? y/n "))
        except:
            mensaje = str(input("El mensaje es correcto? y/n "))
        else:
            if mensaje=="y":
                return True
            else:
                return False
    
def introducir_mensaje():
    global MensajeMasivo
    MensajeMasivo = str(input("Introduzca el Mensaje:\n"))
    if MensajeMasivo:
        print("Mensaje escrito: " + MensajeMasivo)
        mensaje=validarMensaje()
        if not mensaje:
            introducir_mensaje()

def validarImagen():
    imagen = str(input("Desea Enviar Imagen? y/n: "))
    while True:
        try:
            while imagen!="y" and imagen!="n":
                imagen = str(input("Desea Enviar Imagen? y/n: "))
        except:
            imagen = str(input("Desea Enviar Imagen? y/n: "))
        else:
            if imagen=="y":
                return True
            else:
                return False
                
                
def Menu():
   
    Bienvenido(5)
    buscar_base_de_datos()
    time.sleep(1)
    limpiarcmd()
    
    image=validarImagen()
    time.sleep(1)
    limpiarcmd()
    if image:
        introducir_mensaje()
        time.sleep(1)
        limpiarcmd()
        try:
            if BaseDatosMasiva and MensajeMasivo:
                program(BaseDatosMasiva, MensajeMasivo)
        except:
            print("Complete todos los campos antes de comenzar.")
    else:
        introducir_mensaje()
        time.sleep(1)
        limpiarcmd()
        try:
            if BaseDatosMasiva and MensajeMasivo:
                programSinImagenes(BaseDatosMasiva, MensajeMasivo)
        except:
            print("Complete todos los campos antes de comenzar.")
Menu()
