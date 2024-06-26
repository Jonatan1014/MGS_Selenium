import openpyxl
import tkinter as tk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
import easygui as eg
import os
import time
import math
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

global cont
ScanQR=[]
UltimoN=[1]
rutaCarpetaV=[]
extension1 = ["*.xlsx"]
os.system('color 2' if os.name == 'nt' else 'clear')

def limpiarcmd():
    os.system('cls' if os.name == 'nt' else 'clear')

def interfaz():
    hola = """
         __  __                                  _           _____   __  __    _____ 
        |  \/  |                                (_)         / ____| |  \/  |  / ____|
        | \  / |   ___   _ __    ___    __ _     _    ___  | (___   | \  / | | (___  
        | |\/| |  / _ \ | '_ \  / __|  / _` |   | |  / _ \  \___ \  | |\/| |  \___ \ 
        | |  | | |  __/ | | | | \__ \ | (_| |   | | |  __/  ____) | | |  | |  ____) |
        |_|  |_|  \___| |_| |_| |___/  \__,_|   | |  \___| |_____/  |_|  |_| |_____/ 
                                               _/ |                                  
                                              |__/                                   
        by: JSB
        """
    print(hola)

def Bienvenido(duration):
    limpiarcmd()
    chars = ['|', '/', '-', '\\']
    num_chars = len(chars)
    start_time = time.time()
    
    while time.time() - start_time < duration:
        for i in range(num_chars):
            print(f'\rCargando {chars[i]}', end='')
            time.sleep(0.1)
    
    print('\rCarga Completa!   ')
    
    hola = """
         __  __                                  _           _____   __  __    _____ 
        |  \/  |                                (_)         / ____| |  \/  |  / ____|
        | \  / |   ___   _ __    ___    __ _     _    ___  | (___   | \  / | | (___  
        | |\/| |  / _ \ | '_ \  / __|  / _` |   | |  / _ \  \___ \  | |\/| |  \___ \ 
        | |  | | |  __/ | | | | \__ \ | (_| |   | | |  __/  ____) | | |  | |  ____) |
        |_|  |_|  \___| |_| |_| |___/  \__,_|   | |  \___| |_____/  |_|  |_| |_____/ 
                                               _/ |                                  
                                              |__/                                   
        by: JSB
        """
    print(hola)
    chars = ['⣾', '⣽', '⣻', '⢿', '⡿', '⣟', '⣯', '⣷']
    num_chars = len(chars)
    start_time = time.time()
    
    while time.time() - start_time < duration:
        for i in range(num_chars):
            print(f'\rIniciando Programa {chars[i]}', end='')
            time.sleep(0.1)
    
    print('\rIniciado Con Exito!   ')
    time.sleep(2)



def registrarQR(numeroCarpeta):
    
    folder_name = "Google "+str(numeroCarpeta)

    # Obtiene la ruta absoluta de la carpeta actual donde está el programa
    current_directory = os.path.dirname(os.path.abspath(__file__))

    # Ruta completa para la carpeta dentro de la carpeta actual
    folder_path = os.path.join(current_directory, folder_name)

    # Verificar si la carpeta ya existe
    if os.path.exists(folder_path):
        for root, dirs, files in os.walk(folder_path, topdown=False):
            for file in files:
                file_path = os.path.join(root, file)
                os.remove(file_path)
            for dir in dirs:
                dir_path = os.path.join(root, dir)
                os.rmdir(dir_path)

        os.rmdir(folder_path)
        
    # Crea la carpeta
    os.makedirs(folder_path)

    print("Carpeta: "+str(folder_name)+" creada con exito!\nComenzando Registro de QR")
    time.sleep(1)
    chrome_driver_path = "chromedriver.exe"  # Reemplaza con tu ruta
    chrome_service = ChromeService(chrome_driver_path)

    chrome_service.start()

    # Configura las opciones del navegador
    options = webdriver.ChromeOptions()
    # Configura opciones adicionales si es necesario
    # Deshabilita las notificaciones
    options.add_argument("--disable-notifications")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument('user-data-dir='+folder_path)
    # Inicializa el WebDriver utilizando el servicio y las opciones
    browser = webdriver.Chrome(service=chrome_service, options=options)

    # Abre la página web
    browser.get('https://messages.google.com/web/conversations')
    
    # Espera a que el elemento del QR esté presente
    qr_element = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'mw-qr-code img')))

    # Espera a que la imagen del QR cargue
    WebDriverWait(browser, 10).until(lambda driver: qr_element.get_attribute("complete"))
    
    # Boton de recordar inicio de secion
    WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#mat-mdc-slide-toggle-1 > div'))).click()

    # Espera a que cambie el src del elemento del QR (indicando que se escaneó)
    WebDriverWait(browser, 60).until(EC.staleness_of(qr_element))

    # Agrega tu lógica aquí para lo que quieras hacer después de escanear el QR
    # Por ejemplo, puedes imprimir un mensaje indicando que el QR se ha escaneado
    print("El QR se ha escaneado.")
    time.sleep(3)
    browser.quit()
    chrome_service.stop()
    
        
    return (folder_path)

def directorioExcel():
    
    folder_name = "Excel Directorio"

    # Obtiene la ruta absoluta de la carpeta actual donde está el programa
    current_directory = os.path.dirname(os.path.abspath(__file__))

    # Ruta completa para la carpeta dentro de la carpeta actual
    folder_path = os.path.join(current_directory, folder_name)

    # Verificar si la carpeta ya existe
    if os.path.exists(folder_path):
        for root, dirs, files in os.walk(folder_path, topdown=False):
            for file in files:
                file_path = os.path.join(root, file)
                os.remove(file_path)
            for dir in dirs:
                dir_path = os.path.join(root, dir)
                os.rmdir(dir_path)

        os.rmdir(folder_path)
        
    # Crea la carpeta
    os.makedirs(folder_path)

    print("Carpeta: "+str(folder_name)+" creada con exito!")
    time.sleep(1)
    
    return (folder_path)


def dividir_archivo_excel(ruta_archivo, filas_por_grupo, ruta_destino):
    # Cargar el archivo Excel original
    df = pd.read_excel(ruta_archivo, header=None)

    # Calcular el número total de grupos
    total_grupos = math.ceil(len(df) / filas_por_grupo)

    # Iterar sobre los grupos y guardar cada uno en un nuevo archivo Excel
    for i in range(total_grupos):
        inicio = i * filas_por_grupo
        fin = (i + 1) * filas_por_grupo

        grupo_df = df.iloc[inicio:fin]

        # Si es el último grupo, llenar las filas faltantes con "NULL"
        if i == total_grupos - 1:
            filas_faltantes = filas_por_grupo - len(grupo_df)
            for _ in range(filas_faltantes):
                grupo_df = grupo_df.append(['NULL'] * len(df.columns), ignore_index=True)

        # Crear un nuevo archivo Excel para cada grupo
        nuevo_archivo = os.path.join(ruta_destino, f"grupo_{i + 1}.xlsx")
        workbook = Workbook()
        sheet = workbook.active

        for r_idx, row in enumerate(dataframe_to_rows(grupo_df, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Eliminar la primera fila si no es el primer grupo
        if i != 0:
            sheet.delete_rows(1)

        # Guardar el grupo en un nuevo archivo Excel
        workbook.save(filename=nuevo_archivo)

        print(f"Grupo {i + 1} guardado en {nuevo_archivo}")


# Programa sin envio de imagenes
def programSinImagenes(archivo_excel,rutaCarpeta):            

    libro = openpyxl.load_workbook(archivo_excel)
    hoja = libro["Hoja1"]
    chrome_driver_path = "chromedriver.exe"  # Reemplaza con tu ruta
    chrome_service = ChromeService(chrome_driver_path)

    chrome_service.start()

    # Configura las opciones del navegador
    options = webdriver.ChromeOptions()
    # Configura opciones adicionales si es necesario
    # Deshabilita las notificaciones
    options.add_argument("--disable-notifications")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument(r'user-data-dir='+rutaCarpeta)
    # Inicializa el WebDriver utilizando el servicio y las opciones
    browser = webdriver.Chrome(service=chrome_service, options=options)

    # Abre la página web
    browser.get('https://messages.google.com/web/conversations')
    
    cont = 0
    sumN= UltimoN[-1]+1
    error = 0
    for fila in hoja.iter_rows(min_row=UltimoN[-1], values_only=True):  
        Celular = " - ".join(map(str, fila))
        try:
            # Espera hasta que el elemento "loader" no sea visible
            WebDriverWait(browser, 10).until(EC.invisibility_of_element_located((By.ID, 'loader')))
            
            # hacer click en el boton de nuevo mensaje
            # Espera hasta que el elemento sea clickeable
            WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-main-nav > div > mw-fab-link > a > span.mdc-button__label'))).click()
            # hacer click en la entrada de nuevo mensaje
            #browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input').click()

            WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input'))).click()
            

            # agregar numero de telefono
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input').send_keys(Celular)

            # seleccionar el numero ingresado
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > div > mw-contact-selector-button > button').click()
       
            # Espera a que la casilla de texto esté presente y sea visible
            WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'mws-autosize-textarea textarea')))
            
            # Dividir el mensaje en párrafos
            parrafos = MensajeMasivo.split('\n')

            # Escribir cada párrafo en el área de texto
            for i, parrafo in enumerate(parrafos):
                # Escribir el párrafo actual
                browser.find_element(By.CSS_SELECTOR, 'mws-autosize-textarea textarea').send_keys(parrafo)

                # Si no es el último párrafo, enviar con "Shift + Enter" para empezar uno nuevo
                if i < len(parrafos) - 1:
                    ActionChains(browser).key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT).perform()

            # Enviar el último párrafo
            browser.find_element(By.CSS_SELECTOR, 'mws-autosize-textarea textarea').send_keys(Keys.ENTER)
            
            cont=cont+1
            print(cont)
            if cont>249:
                UltimoN.append(sumN+1)
                break
                    
        except StaleElementReferenceException as e:
            print("Error: Elemento de página obsoleto. ", str(e))
        except ElementClickInterceptedException as e:
            print("Error: Intercepción de clic en el elemento. ", str(e))
        except Exception as e:
            print("Ocurrió un error inesperado: ", str(e))
            error+=1
            if error>2:
                UltimoN.append(sumN-3)
                break
            
    
    # Cierra el navegador y detiene el servicio
    time.sleep(2)
    ScanQR.append(1)
    browser.quit()
    chrome_service.stop()
    print("Envio completado")
    libro.close()

# Programa con envio de imagenes
def program(archivo_excel,rutaCarpeta):            

    libro = openpyxl.load_workbook(archivo_excel)
    hoja = libro["Hoja1"]
    chrome_driver_path = "chromedriver.exe"  # Reemplaza con tu ruta
    chrome_service = ChromeService(chrome_driver_path)

    chrome_service.start()

    # Configura las opciones del navegador
    options = webdriver.ChromeOptions()
    # Configura opciones adicionales si es necesario
    # Deshabilita las notificaciones
    options.add_argument("--disable-notifications")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument('user-data-dir='+rutaCarpeta)
    # Inicializa el WebDriver utilizando el servicio y las opciones
    browser = webdriver.Chrome(service=chrome_service, options=options)

    # Abre la página web
    browser.get('https://messages.google.com/web/conversations')
    
    cont = 0
    sumN= UltimoN[-1]+1
    error = 0
    for fila in hoja.iter_rows(min_row=UltimoN[-1], values_only=True):  
        Celular = " - ".join(map(str, fila))
        try:
            # Espera hasta que el elemento "loader" no sea visible
            WebDriverWait(browser, 10).until(EC.invisibility_of_element_located((By.ID, 'loader')))
            
            # hacer click en el boton de nuevo mensaje
            # Espera hasta que el elemento sea clickeable
            WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-main-nav > div > mw-fab-link > a > span.mdc-button__label'))).click()
            # hacer click en la entrada de nuevo mensaje
            #browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input').click()

            WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input'))).click()
            

            # agregar numero de telefono
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > mw-new-conversation-sub-header > div > div.input-container > mw-contact-chips-input > div > div > input').send_keys(Celular)

            # seleccionar el numero ingresado
            browser.find_element(By.CSS_SELECTOR, 'body > mw-app > mw-bootstrap > div > main > mw-main-container > div > mw-new-conversation-container > div > mw-contact-selector-button > button').click()
       
            # Espera a que la casilla de texto esté presente y sea visible
            WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'mws-autosize-textarea textarea')))
            
            # Dividir el mensaje en párrafos
            parrafos = MensajeMasivo.split('\n')

            # Escribir cada párrafo en el área de texto
            for i, parrafo in enumerate(parrafos):
                # Escribir el párrafo actual
                browser.find_element(By.CSS_SELECTOR, 'mws-autosize-textarea textarea').send_keys(parrafo)

                # Si no es el último párrafo, enviar con "Shift + Enter" para empezar uno nuevo
                if i < len(parrafos) - 1:
                    ActionChains(browser).key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT).perform()

            # Enviar el último párrafo
            browser.find_element(By.CSS_SELECTOR, 'mws-autosize-textarea textarea').send_keys(Keys.ENTER)
            
            # Pega el texto en el campo de entrada usando la combinación de teclas "Control + V"
            ActionChains(browser).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
            
            # Enviar imagen con Enter
            browser.find_element(By.CSS_SELECTOR, 'mws-autosize-textarea textarea').send_keys(Keys.ENTER)
            
            cont=cont+1
            print(cont)
            if cont>249:
                UltimoN.append(sumN+1)
                break
                    
        except StaleElementReferenceException as e:
            print("Error: Elemento de página obsoleto. ", str(e))
        except ElementClickInterceptedException as e:
            print("Error: Intercepción de clic en el elemento. ", str(e))
        except Exception as e:
            print("Ocurrió un error inesperado: ", str(e))
            error+=1
            if error>2:
                UltimoN.append(sumN-3)
                break
    
    
    
    # Cierra el navegador y detiene el servicio
    ScanQR.append(3)
    browser.quit()
    chrome_service.stop()
    print("Envio completado")
    libro.close()



def contar_filas(nombre_hoja,archivo_excel):
    # Carga el archivo de Excel
    libro = openpyxl.load_workbook(archivo_excel)
    
    # Accede a la hoja específica
    hoja = libro[nombre_hoja]
    
    # Cuenta las filas en la hoja
    numero_filas = hoja.max_row
    
    return numero_filas

def buscar_base_de_datos():
    global BaseDatosMasiva
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal
    print("Buscar Base de Datos")
    BaseDatosMasiva = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if BaseDatosMasiva:
        print("Base de Datos seleccionada: " + BaseDatosMasiva)
        return True, BaseDatosMasiva
    else:
        salir=str(input("Base de Datos no Seleccionada\n¿Desea salir? y/n\n"))
        limpiarcmd()
        while True:
            try:
                while salir!="y" and salir!="n":
                    salir=str(input("Opcion Invalida\n¿Desea salir? y/n\n"))
                    limpiarcmd()
            except:
                salir=str(input("Base de Datos no Seleccionada\n¿Desea salir? y/n\n"))
            else:
                if salir=="y":
                    break
                else:
                    buscar_base_de_datos()

def introducir_mensaje():
    global MensajeMasivo
    MensajeMasivo = eg.codebox(msg='Entrada de fuente',title='Control: codebox')
    if MensajeMasivo:
        print("Mensaje escrito: " + MensajeMasivo)
        mensaje=validarMensaje()
        if not mensaje:
            introducir_mensaje()

def validarMensaje():
    mensaje = str(input("El mensaje es correcto? y/n "))
    while True:
        try:
            while mensaje!="y" and mensaje!="n":
                mensaje = str(input("El mensaje es correcto? y/n "))
        except:
            mensaje = str(input("El mensaje es correcto? y/n "))
        else:
            if mensaje=="y":
                return True
            else:
                return False

def validarImagen():
    imagen = str(input("Desea Enviar Imagen? y/n: "))
    while True:
        try:
            while imagen!="y" and imagen!="n":
                imagen = str(input("Desea Enviar Imagen? y/n: "))
        except:
            imagen = str(input("Desea Enviar Imagen? y/n: "))
        else:
            if imagen=="y":
                return True
            else:
                return False
                
def Menu():
    Bienvenido(1)
    ruta_archivo_original,validarBD=buscar_base_de_datos()
    if validarBD:
        time.sleep(1)
        limpiarcmd()
        interfaz()
        ruta_destino=directorioExcel()
        dividir_archivo_excel(ruta_archivo_original, 250, ruta_destino)
        time.sleep(1)
        limpiarcmd()
        interfaz()

        image=validarImagen()
        time.sleep(1)
        limpiarcmd()
        interfaz()

        if image:
            introducir_mensaje()
            time.sleep(1)
            limpiarcmd()
            interfaz()
            try:
                if BaseDatosMasiva and MensajeMasivo:
                    numeroQR=int(input("Introduzca la cantidad de QR"))
                    rutaCarpetaV = [""] * (numeroQR + 1)  # Inicializa la lista con elementos vacíos
                    for numeroCarpeta in range(1,numeroQR+1):
                            rutaCarpeta=registrarQR(numeroCarpeta)
                            rutaCarpetaV[numeroCarpeta]=rutaCarpeta
                            
                    fin = contar_filas("Hoja1", BaseDatosMasiva)
                    intervalos= fin/250
                    
                    numero =0
                    for i in range(0,int(math.ceil(intervalos))) :
                        numero = i % 2 + 1
                        program(BaseDatosMasiva,rutaCarpetaV[numero])

                    
                    print("Trabajo terminado corrrectamente")
                    
            except:
                print("Complete todos los campos antes de comenzar.")
        else:
            introducir_mensaje()
            time.sleep(1)
            limpiarcmd()
            interfaz()
            if BaseDatosMasiva and MensajeMasivo:
                numeroQR=int(input("Introduzca la cantidad de QR"))
                rutaCarpetaV = [""] * (numeroQR + 1)  # Inicializa la lista con elementos vacíos
                for numeroCarpeta in range(1,numeroQR+1):
                        rutaCarpeta=registrarQR(numeroCarpeta)
                        rutaCarpetaV[numeroCarpeta]=rutaCarpeta
                        
                fin = contar_filas("Hoja1", BaseDatosMasiva)
                intervalos= fin/250
                
                numero =0
                for i in range(0,int(math.ceil(intervalos))) :
                    numero = i % 2 + 1
                    programSinImagenes(BaseDatosMasiva,rutaCarpetaV[numero])

                
                print("Trabajo terminado corrrectamente")
            else:
        
                print("Complete todos los campos antes de comenzar.")
    else:
        print("Adios!")


Menu()
